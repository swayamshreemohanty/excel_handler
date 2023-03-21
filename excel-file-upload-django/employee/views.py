import getpass
import os
from pyexpat.errors import messages
from django.http import JsonResponse
from django.shortcuts import render
from django.shortcuts import render, redirect
import openpyxl
import pandas as pd  
from employee.forms import EmployeeForm  
from employee.models import Employee, User  
from django.middleware import csrf
#Excel file name
excel_file_name="demo_data.xlsx"



# Static
def login(request):        
    request=request.POST
    username=request['username']
    password=request['password']
    isSame=(username=='admin' and password=='admin')
       
    if isSame:
        return JsonResponse({"status":True, "message":"User logined successfully"}, status=200)
    else:
        return JsonResponse({"status":False, "message": "Username and Password not matched"}, status=400)

def handle_uploaded_file(f, path):
    with open(path, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
    
def storefile(request):
    excel_file = request.FILES["excel_file"]
    #make the final directory path
    file_path=os.path.join(getExcelDirectory()+excel_file_name)
        #store the file
    handle_uploaded_file(excel_file,file_path)
        
    return redirect('/show')

def getExcelDirectory():
    #get the machine user name
    username=getpass.getuser();
    dir_path=os.path.join('C:/Users/{}/Desktop/Excelfile/'.format(username)) 
    
    # create directory if it does not exist
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    return dir_path

def fetchEmployeeFromExcel():
    try:
        file_path=os.path.join(getExcelDirectory()+excel_file_name)
    # Load the Excel file
        workbook = openpyxl.load_workbook(file_path)
            
        emp_list=[]
    # Select the first worksheet
        worksheet = workbook.active
        for row in worksheet.iter_rows(min_row=2):
            row_dict = {}
            
            for cell in row:
                row_dict[cell.column_letter] = cell.value   
            emp_id=row_dict['A']
            application=row_dict['B']
            budget=row_dict['C']
            valuation=row_dict['D']
              
            emp_data=Employee(eid=emp_id,application=application,valuation=valuation,budget=budget)
            emp_list.append(emp_data)
        return emp_list
    except Exception:
        return []

def editEmployee(employee:Employee,id:int):
    file_path=os.path.join(getExcelDirectory()+excel_file_name)        
    # open the Excel file and select the sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
            
    # find the row with the matching ID in the first column
    # id_found=False
    for row in sheet.iter_rows(min_row=2, values_only=True):        
        if  row[0] == id:
            # update the row with the new data
            row_data = list(row)
            row_data[1]="{}".format(employee.application)
            row_data[2]="{}".format(employee.budget)
            row_data[3]="{}".format(employee.valuation)
            for col_idx, cell in enumerate(row_data, start=1):
                print(cell)
                sheet.cell(row=row[0]+1, column=col_idx, value=cell)
                wb.save(file_path)     
            return True
    return False
    
def readLastId(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the worksheet
    worksheet = workbook.active

    # Find the last row in the column
    last_row = worksheet.max_row

    # Read the value of the last cell in the column
    last_cell_value = worksheet.cell(row=last_row, column=1).value
    
    return last_cell_value

def emp(request):  
    if request.method == "POST":  
        # form = EmployeeForm(request.POST)
        try:  
            application=request.POST['application']
            valuation=request.POST['valuation']
            budget=request.POST['budget']  
            new_emp=Employee(application=application,valuation=valuation,budget=budget)
            addEmployee(new_emp)  
            return redirect('/show')
        except Exception:
            pass
         
    else:  
        form = EmployeeForm()  
    return render(request,'index.html',{'form':form})

def addEmployee(employee:Employee):
      
    #make the final directory path
    file_path=os.path.join(getExcelDirectory()+excel_file_name)
    loadedExcelFile = pd.read_excel(file_path)
            
    #Insert
    last_id=readLastId(file_path)
    #create new row
    new_row={'ID':last_id+1,'APLLICATION':employee.application,'BUDGET':"${}".format(employee.budget),'VALUATION':"${}".format(employee.valuation)}
    #append the new row to data frame
    loadedExcelFile = loadedExcelFile.append(new_row, ignore_index=True)
            
    # save modified Excel file
    loadedExcelFile.to_excel(file_path,index=False)
            
    employee.save()

def getEmployeeById(id):
    file_path=os.path.join(getExcelDirectory()+excel_file_name)
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
            
    # Select the first worksheet
    worksheet = workbook.active
                
    # Loop through the rows and search for a matching ID
    for row in worksheet.iter_rows(min_row=2):
        if  row[0].value == id:
            # If a matching ID is found, create a dictionary representing the row
            row_dict = {}
                        
            for cell in row:
                row_dict[cell.column_letter] = cell.value
            emp_id=row_dict['A']
            application=row_dict['B']
            budget=row_dict['C']  
            valuation=row_dict['D']
            emp_data=Employee(eid=emp_id,application=application,valuation=valuation,budget=budget)
            return emp_data

def show(request):     
    employees = fetchEmployeeFromExcel()  
    # employees = Employee.objects.all()
    if len(employees)==0:
        return render(request,"upload.html")   
    else:
        return render(request,"show.html",{'employees':employees})  

def edit(request, id):  
    # employee = Employee.objects.get(id=id)  
    employee= getEmployeeById(id)
    return render(request,'edit.html', {'employee':employee}) 

def update(request, id): 
    # employee= getEmployeeById(id)
    # form = EmployeeForm(request.POST, instance = employee)  
    # if form.is_valid():  
    #     form.save()  
    try:
        application=request.POST['application']
        valuation=request.POST['valuation']
        budget=request.POST['budget'] 
        updatedData=Employee(eid=id,application=application,valuation=valuation,budget=budget)
        editEmployee(updatedData,id)
        return redirect("/show")  
    except Exception: 
        # return render(request, 'edit.html', {'employee': employee}) 
        return redirect("/show")  


#API methods
def upload_excel(request):
    excel_file = request.FILES["excel_file"]
    #make the final directory path
    file_path=os.path.join(getExcelDirectory()+excel_file_name)
    #store the file
    handle_uploaded_file(excel_file,file_path)  
    return JsonResponse({"status":True, "message":"File uploaded successfully"}, status=200)

def add_emp(request):  
    if request.method == "POST": 
        try:  
            application=request.POST['application']
            valuation=request.POST['valuation']
            budget=request.POST['budget']  
            new_emp=Employee(application=application,valuation=valuation,budget=budget)            
            addEmployee(new_emp)  
            return JsonResponse({"status":True, "message":"User added successfully"}, status=200)
        except Exception as e:
            return JsonResponse({"status":True, "message":"Unable to add user"}, status=400)

    else:  
        return JsonResponse({"status":False, "message":"Unable to add user"}, status=400)

def update_emp(request): 
    try:
        application=request.POST['application']
        valuation=request.POST['valuation']
        budget=request.POST['budget'] 
        updatedData=Employee(eid=request.POST['id'],application=application,valuation=valuation,budget=budget)
        isUpdated=editEmployee(updatedData,int(request.POST['id']))
        if isUpdated:
            return JsonResponse({"status":True, "message":"User updated successfully"}, status=200)
        else:
            return JsonResponse({"status":False, "message":"Unable to update User"}, status=400)
    except Exception as e: 
        return JsonResponse({"status":False, "message": str(e)}, status=400)
        # return redirect("/show")  

def fetch_all_emp(request):
    try:
        employees = fetchEmployeeFromExcel()  
        if len(employees)==0:
            error = {'message': "No data found", 'status': 400}
            return JsonResponse(error, status=400)
        else:
                # Convert the array of objects to a list of dictionaries
            my_dicts = []
            for obj in employees:
                my_dict = {"id": obj.eid,"application": obj.application,"valuation": obj.valuation,"budget": obj.budget}
                my_dicts.append(my_dict)

            data = {'data': my_dicts, 'status': 200}
            return JsonResponse(data, status=200, safe=False)
    
    except Exception as e:
        error = {'message': str(e), 'status': 400}
        return JsonResponse(error, status=400)

def fetch_emp_by_id(request):  
    try:
        # Get the ID from the query parameter       
        row_id = request.GET.get('id')
        employee= getEmployeeById(int(row_id))
        my_dict = {"id": employee.eid, "application": employee.application,"valuation": employee.valuation,"budget": employee.budget}
        data = {'data': my_dict, 'status': 200}
        return JsonResponse(data, status=200, safe=False)
    except Exception as e:
        return JsonResponse({"status":False, "message":"No data found"}, status=400)


# def register(request):
#     if request.method == "POST":   
#         errors = User.objects.register_validator(request.POST) 
#         if len(errors):
#             for key, value in errors.items():
#                 messages.add_message(request, messages.ERROR, value, extra_tags='register')
#             # return redirect('/')
#             return JsonResponse({"status":False, "message":value}, status=200)
#         else:
#             pw_hash =  request.POST['password']
#             # crypt.hashpw(request.POST['password'].encode(), scrypt.gensalt())
#             user = User.objects.create(first_name=request.POST['first_name'], last_name=request.POST['last_name'], email=request.POST['email'], password=pw_hash)
#             request.session['user_id'] = user.id
#             # return redirect("/success")
#             return JsonResponse({"status":True, "message":"User registered successfully"}, status=200)      
#     else:
#         return redirect("/")

# def login(request):
#     if request.method == "POST":        
#         errors = User.objects.login_validator(request.POST)
#         if len(errors):
#             for key, value in errors.items():
#                 messages.add_message(request, messages.ERROR, value, extra_tags='login')
#             return redirect('/')
#         else:
#             user = User.objects.get(email=request.POST['email'])
#             request.session['user_id'] = user.id
#             return redirect("/show")
    



def get_csrf_token(request):
    if request.method=='GET':
        token = csrf.get_token(request)
        return JsonResponse({"status":True, "token":token}, status=200)
    else:
        return JsonResponse({"status":False, "message":"Unable to get token"}, status=400)


def destroy(request, id):  
    employee = Employee.objects.get(id=id)  
    employee.delete()  
    return redirect("/show")