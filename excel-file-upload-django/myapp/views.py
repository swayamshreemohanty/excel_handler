from getpass import getpass
import json
import os
from django.shortcuts import render
import openpyxl
from .models import Data
from django.http import HttpResponse, JsonResponse
from datetime import datetime
import pandas as pd
import getpass
from django.middleware import csrf

#Excel file name
excel_file_name="employee_data.xlsx"

def handle_uploaded_file(f, path):
    with open(path, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
            
def getExcelDirectory():
    #get the machine user name
    username=getpass.getuser();
    dir_path=os.path.join('C:/Users/{}/Desktop/Excelfile/'.format(username)) 
    
    # create directory if it does not exist
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    return dir_path

def index(request):
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size
                
        #Store the file in the system folder
    
        #make the final directory path
        file_path=os.path.join(getExcelDirectory()+excel_file_name)
        #store the file
        handle_uploaded_file(excel_file,file_path)
        
        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

        return render(request, 'myapp/index.html', {"excel_data":excel_data})



def readLastId(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the worksheet
    worksheet = workbook.active

    # Find the last row in the column
    last_row = worksheet.max_row

    # Read the value of the last cell in the column
    last_cell_value = worksheet.cell(row=last_row, column=1).value
    
    return last_cell_value


def add_emp(request):
    if request.method=='POST':
        try:
            application=request.POST['application']
            valuation=int(request.POST['valuation'])
            budget=int(request.POST['budget'])        
            new_emp=Data(application=application,valuation=valuation,budget=budget)
            #make the final directory path
            file_path=os.path.join(getExcelDirectory()+excel_file_name)
            loadedExcelFile = pd.read_excel(file_path)
            
            #Insert
            last_id=readLastId(file_path)
            #create new row
            new_row={'ID':last_id+1,'APLLICATION':new_emp.application,'BUDGET':"${}".format(new_emp.budget),'VALUATION':"${}".format(new_emp.valuation)}
            #append the new row to data frame
            loadedExcelFile = loadedExcelFile.append(new_row, ignore_index=True)
            
            # save modified Excel file
            loadedExcelFile.to_excel(file_path,index=False)
            
            new_emp.save()
            return JsonResponse({"status":True, "message":'Date added sucessfully'}, status=200)
        
        except Exception as e:
            error = {'error': str(e), 'status': 400}
            return JsonResponse(error, status=400)
            
            
    elif request.method=='GET':
        return render(request,"myapp/add_emp.html")
    else:
        return HttpResponse("An exception ocured")
    
def update_emp(request):
    if request.method=='POST':
        try:
            row_id = request.POST.get('id', None)
            if row_id!="None": 
                application=request.POST['application']
                valuation=int(request.POST['valuation'])
                budget=int(request.POST['budget'])        
                new_emp=Data(application=application,valuation=valuation,budget=budget)
                #make the final directory path
                file_path=os.path.join(getExcelDirectory()+excel_file_name)
                
                # open the Excel file and select the sheet
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active
                
            
                # find the row with the matching ID in the first column
                id_found=False
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if str(row[0]) == row_id:
                        id_found=True
                        # update the row with the new data
                        row_data = list(row)
                        row_data[1]=new_emp.application
                        row_data[2]="${}".format(new_emp.budget)
                        row_data[3]="${}".format(new_emp.valuation)
                        for col_idx, cell in enumerate(row_data, start=1):
                            sheet.cell(row=row[0]+1, column=col_idx, value=cell)
                        break
                if id_found:
                    # save the changes to the Excel file
                    wb.save(file_path)
                    return JsonResponse({"status":True, "message":'Date updated sucessfully'}, status=200)
                else:
                    error = {'error': "Id missing", 'status': 400}
                    return JsonResponse(error, status=400)
            else:
                error = {'error': "Id missing", 'status': 400}
                return JsonResponse(error, status=400)
            
        except Exception as e:
            error = {'error': str(e), 'status': 400}
            return JsonResponse(error, status=400)
            
            
    elif request.method=='GET':
        return render(request,"myapp/add_emp.html")
    else:
        return HttpResponse("An exception ocured")
    

def fetch_emp(request):
    try:
        if request.method=='GET':
            # Get the ID from the query parameter       
            row_id = request.GET.get('id', None)
            if row_id!="None":
                # #make the final directory path
                # loadedExcelFile = pd.read_excel(file_path)
                
                file_path=os.path.join(getExcelDirectory()+excel_file_name)
                # Load the Excel file
                workbook = openpyxl.load_workbook(file_path)
            
                # Select the first worksheet
                worksheet = workbook.active
                
                # Loop through the rows and search for a matching ID
                for row in worksheet.iter_rows(min_row=2):
                    if str(row[0].value) == row_id:
                        # If a matching ID is found, create a dictionary representing the row
                        row_dict = {}
                        
                        for cell in row:
                            row_dict[cell.column_letter] = cell.value

                        # Convert the dictionary to a JSON string
                        json_data = json.dumps(row_dict)
                        
                        # Return the row as a JSON response
                        return HttpResponse(json_data,content_type='application/json')
                    
                error = {'error': "Id missing", 'status': 400}
                return JsonResponse(error, status=400)
            
            else:
                error = {'error': "Id missing", 'status': 400}
                return JsonResponse(error, status=400)
       
        error = {'error': "GET method required", 'status': 400}
        return JsonResponse(error, status=400)
    
    except Exception as e:
        error = {'error': str(e), 'status': 400}
        return JsonResponse(error, status=400)


def get_csrf_token(request):
    if request.method=='GET':
        token = csrf.get_token(request)
        json_data = json.dumps({"Token":token})
        return HttpResponse(json_data,content_type='application/json')
    
    else:
        return HttpResponse("An exception ocured")


